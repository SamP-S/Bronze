import os
import pandas as pd
import numpy as np
import argparse
import csv
from openpyxl import load_workbook
import importlib.resources as pkg_resources


# loads template.xlsx file and returns sheets as dictionary of pd.DataFrames or empty dictionary if failed
### TODO:
# copy template.xlsx to output path then write new qto sheets to it
# means existing formatting will be preserved instead of getting trunced by dataframe
# returns template workbook
# TODO: Package template.xlsx with scs package to ensure it is kept with the sourcecode
def get_template_wb():    
    
    # TODO: hack solution to not have to recreate formulas
    # create template sheets inside of the template.xlsx
    # make it 100s of lines of usable space, fill as needed then delete unused rows
    # leaving a few as a buffer. Allows for customising the template without having
    # to program it in python.
    # makes it a fuck to easier but does force template.xlsx to exist else no formatting
    # which makes it practically useless
    
    try:
        with pkg_resources.open_binary('scs.converters', 'template.xlsx') as file:
            workbook = load_workbook(file)
            return workbook
    except Exception as err:
        print(f"WARNING: Couldn't open template, skipping...\n{err}")
        return None

    

# converts maxcut csv to xlsx quote
# returns path to quote or None if failed
def maxcut_to_quote(input_path, output_path=None):
    if input_path is None:
        print(f"ERROR: Input path is None.\n{input_path}")
        return None
    
    if not os.path.exists(input_path):
        print(f"ERROR: Can't find input file.\n{input_path}")
        return None
        
    if not input_path.endswith(".csv"):
        print(f"ERROR: input file is not CSV file.\n{input_path}")
        return None
    
    
    if output_path is None:
        print(f"INFO: Output path is None. Using default.\n{output_path}")
        filename_ext = os.path.splitext(input_path)
        output_path = filename_ext[0] + "_quote.xlsx"
    
    if os.path.exists(output_path):
        print(f"ERROR: Output file already exists. Refusing to overwrite.\n{output_path}")
        return None
    
    output_dir = os.path.dirname(output_path)
    if not os.path.exists(output_dir) and len(output_dir) > 0:
        os.makedirs(output_dir, exist_ok=True)
        print(f"INFO: Output path parent directory not found. Creating output directory.\n{output_dir}")

    # attempt to load csv
    df = None
    try:
        # read first line to get delimiter
        with open(input_path, 'r') as file:
            line = file.readline()
            if (line.startswith("Sep=")):
                sep = line.strip().split('=')[1]
                df = pd.read_csv(input_path, sep=sep, skiprows=1)
            else:
                df = pd.read_csv(input_path, sep=",")
        
    except Exception as err:
        return f"ERROR: Couldn't open file.\n{input_path}\n {err}\nPlease close the file if you have it opened."
    
    # parse df
    ### TODO: condider other columns to include/use for categorization/subcategorization for generating main stone sheet
    ### TODO: currently skipping grouping & edging atm edge_N, Import ID, Parent ID
    print(f"columns = {df.columns}")
    df = df[["Type", "Name", "Length", "Width", "Quantity", "Notes", "Material"]]
    df = df[~df["Type"].isin(["Input Labour", "Input Hardware", "Input Edging", "Input Group"])]

    # convert dimension strings to floats
    df["Length"] = (df["Length"].str[:-3]).astype(float)
    df["Width"] = (df["Width"].str[:-3]).astype(float)
    # replace NaN notes with empty string
    df["Notes"] = df["Notes"].fillna("")
    
    materials = df["Material"].unique()
    print(f"INFO: Found material set:\n{materials}")
    print(f"INFO: df shape = {df.shape}")
    print(f"INFO: df columns = {df.columns}")
    print(f"INFO: df:\n{df}")
    
    for material in materials:
        print(f"INFO: Processing material: {material}")
        material_df = df[df["Material"] == material]
        print(f"INFO: material_df shape = {material_df.shape}")
        
        
    wb = get_template_wb()
    print(f"INFO: workbook:\n{wb}")
    print(f"INFO: sheetnames:\n{wb.sheetnames}")
    print(f"INFO: workbook path:\n{wb.path}")
    
    return None


### IMPORTANT IMPLEMENTATION NOTES:
# "./template.xlsx" hold "Front Cover", "Quote" and "Allowances" sheets to copy into generated quote.
# It is assumed all relevant maxcut input items are contained in a single csv file.
# Input items will be seperated by material into respective QTO sheets.
# Groups should be used to determine which room a material belongs to, if no groups found, all materials will be assumed to be in a single unspecified room.
# All input items of type "labour" and "hardware" will be ignored.
# TODO: Accept edging
# TODO: Support .mc3 files

def main():

    parser = argparse.ArgumentParser(
        prog="scs_quote_to_maxcut",
        description="convert quote sheet from .xlsx to maxcut .csv"
    )
    # Required arguments
    parser.add_argument("filepath")
    # Optional arguments
    parser.add_argument("-o", "--output_path", help="output filepath")
    # Compile args
    args = parser.parse_args()

    # convert maxcut csv to quote xlsx
    quote_path = maxcut_to_quote(args.filepath, args.output_path)
    if quote_path is None:
        print(f"INFO: Conversion successful. Output written to:\n{quote_path}")
    else:
        print(f"ERROR: Conversion failed.\n{quote_path}")
        exit(1)

if __name__ == "__main__":
    main()